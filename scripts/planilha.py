import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
import math
from io import BytesIO

def run():
    st.write("""
    # 📊 Criar planilha de resumo RTI
    #### por enquanto apenas NBR 5410...""")

    uploaded_file = st.file_uploader("Envie seu arquivo Excel", type=["xlsx"])

    if uploaded_file is not None: #pandas
        df = pd.read_excel(uploaded_file)
        if "Dados Cliente" in df.columns:
            df["Área/Subestação"] = df["Dados Cliente"].str.extract(r"Área\/Subestação:\s*([^\n\r]+)")
            df["Equipamento"] = df["Dados Cliente"].str.extract(r"Equipamento:\s*([^\n\r]+)")

            df = df.drop(df.columns[:4], axis=1)
            df = df[["Área/Subestação", "Equipamento"] + [c for c in df.columns if c not in ["Área/Subestação", "Equipamento"]]]
            selecionadas = df.columns[:2].tolist() + df.columns[3::4][:-1].tolist()
            df = df[selecionadas]

        #openpyxl
        buffer = BytesIO()
        df.to_excel(buffer, index=False)
        buffer.seek(0)

        wb = openpyxl.load_workbook(buffer)
        ws = wb.active

        for col in ["A", "B"]:
            for cell in ws[col]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
            ws.column_dimensions[col].width = 20

        for col in [chr(c) for c in range(ord("C"), ord("U") + 1)]:
            ws.column_dimensions[col].width = 16

        ws.row_dimensions[1].height = 75
        for cell in ws[1]:
            cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")

        last_row = ws.max_row
        for r in range(2, last_row + 1):
            ws.row_dimensions[r].height = 30

        ws.column_dimensions["V"].width = 9
        for cell in ws["V"]:
            
            cell.font = Font(bold=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")

        dv = DataValidation(type="list", formula1='"Conforme,Não conforme,Não se aplica"', allow_blank=True)
        rng = f"C2:U{last_row}"
        dv.add(rng)
        ws.add_data_validation(dv)

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        blue_fill = PatternFill(start_color="91CDDC", end_color="91CDDC", fill_type="solid")

        ws.conditional_formatting.add(rng, FormulaRule(formula=["C2=\"Conforme\""], fill=green_fill))
        ws.conditional_formatting.add(rng,
                                      FormulaRule(formula=["C2=\"Não conforme\""], fill=red_fill, font=Font(bold=True)))
        ws.conditional_formatting.add(rng,
                                      FormulaRule(formula=["C2=\"Não se aplica\""], fill=blue_fill, font=Font(bold=True)))

        first_col = ws["C1"].column
        last_col = ws["U1"].column
        totals_row = last_row + 1
        totals_col = last_col + 1

        ws.cell(row=totals_row, column=2, value="Total não conforme")

        for c in range(first_col, last_col + 1):
            ws.cell(row=totals_row, column=c,
                    value=f'=COUNTIF({ws.cell(2, c).coordinate}:{ws.cell(last_row, c).coordinate},"não conforme")')

        ws.cell(row=1, column=totals_col, value="Total não conforme")

        for r in range(2, last_row + 1):
            ws.cell(row=r, column=totals_col,
                    value=f'=COUNTIF({ws.cell(r, first_col).coordinate}:{ws.cell(r, last_col).coordinate},"não conforme")')
        last_row = ws.max_row
        ws.row_dimensions[last_row].height = 15
        ws.font = Font(bold=True)
        for cell in ws[last_row]:
            cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")

        for c in range(first_col, totals_col + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
            ws.cell(row=totals_row, column=c).font = Font(bold=True)

        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in ws.iter_rows(min_row=1, max_row=totals_row, min_col=1, max_col=totals_col):
            for cell in row:
                cell.border = border

        output = BytesIO()
        wb.save(output)
        st.download_button(
            label="⬇️ Baixar Excel formatado",
            data=output.getvalue(),
            file_name="resultado_formatado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
